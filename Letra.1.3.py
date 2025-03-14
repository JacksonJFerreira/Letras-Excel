import os
import threading
import customtkinter as ctk
from tkinter import StringVar, filedialog, messagebox
from openpyxl import load_workbook

# Definir aparência do CustomTkinter
ctk.set_appearance_mode("System")  # Modos: "System" (padrão), "Dark", "Light"
ctk.set_default_color_theme("blue")  # Temas: "blue" (padrão), "green", "dark-blue"

def escolher_pasta():
    """Abre uma caixa de diálogo para o usuário escolher a pasta que contém os arquivos Excel."""
    pasta = filedialog.askdirectory()
    if pasta:
        pasta_var.set(pasta)
        label_pasta.configure(text=pasta)

def processar_arquivos(pasta, coluna):
    """Percorrer todos os arquivos .xlsx na pasta indicada, abrir cada planilha (em todas as abas) e transformar o conteúdo das células da coluna informada para letras maiúsculas."""
    # Filtrar os arquivos que possuem extensão .xlsx
    arquivos = [f for f in os.listdir(pasta) if f.lower().endswith('.xlsx')]
    total = len(arquivos)
    if total == 0:
        messagebox.showwarning("Aviso", "Nenhum arquivo Excel (.xlsx) encontrado na pasta.")
        progress_bar.set(0)
        return

    progress_bar.configure(determinate_speed=1/total)
    cont = 0

    for arquivo in arquivos:
        caminho = os.path.join(pasta, arquivo)
        try:
            wb = load_workbook(caminho)
            # Itera por todas as abas da planilha
            for ws in wb.worksheets:
                try:
                    # Percorre todas as células da coluna informada
                    for celula in ws[coluna]:
                        if celula.value is not None and isinstance(celula.value, str):
                            celula.value = celula.value.upper()
                except Exception:
                    # Caso a coluna não exista na aba, ignora
                    pass
            wb.save(caminho)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao processar o arquivo {arquivo}.\nErro: {e}")
        cont += 1
        progress_bar.set(cont/total)
        status_label.configure(text=f"Processando: {cont}/{total} arquivos")
        root.update_idletasks()

    progress_bar.set(1)
    status_label.configure(text="Processamento concluído!")
    messagebox.showinfo("Concluído", "Processo de transformação concluído.")
    progress_bar.set(0)

def iniciar_processamento():
    pasta = pasta_var.get()
    coluna = coluna_var.get().upper()
    if not pasta or not coluna:
        messagebox.showwarning("Aviso", "Por favor, selecione uma pasta e informe a coluna.")
        return
    
    status_label.configure(text="Iniciando processamento...")
    btn_iniciar.configure(state="disabled")
    btn_escolher_pasta.configure(state="disabled")
    
    threading.Thread(target=lambda: [
        processar_arquivos(pasta, coluna), 
        root.after(0, lambda: [
            btn_iniciar.configure(state="normal"),
            btn_escolher_pasta.configure(state="normal")
        ])
    ]).start()

# Criação da interface gráfica
root = ctk.CTk()
root.title("Conversor para Maiúsculas - Excel")
root.geometry("500x400")
root.resizable(False, False)

# Criando um frame principal para organizar os widgets
main_frame = ctk.CTkFrame(root)
main_frame.pack(padx=20, pady=20, fill="both", expand=True)

# Título
title_label = ctk.CTkLabel(main_frame, text="Conversor para Maiúsculas", font=ctk.CTkFont(size=20, weight="bold"))
title_label.pack(pady=10)

# Descrição
desc_label = ctk.CTkLabel(main_frame, text="Converte para maiúsculas todos os dados da coluna selecionada\nem todos os arquivos Excel (.xlsx) da pasta escolhida.")
desc_label.pack(pady=(0, 15))

# Frame para seleção de pasta
pasta_frame = ctk.CTkFrame(main_frame)
pasta_frame.pack(fill="x", padx=10, pady=5)

pasta_label = ctk.CTkLabel(pasta_frame, text="Pasta:")
pasta_label.pack(side="left", padx=5)

pasta_var = StringVar()
label_pasta = ctk.CTkLabel(pasta_frame, text="Nenhuma pasta selecionada", width=250, anchor="w")
label_pasta.pack(side="left", padx=10, fill="x", expand=True)

btn_escolher_pasta = ctk.CTkButton(pasta_frame, text="Escolher", command=escolher_pasta, width=100)
btn_escolher_pasta.pack(side="right", padx=5)

# Frame para entrada da coluna
coluna_frame = ctk.CTkFrame(main_frame)
coluna_frame.pack(fill="x", padx=10, pady=5)

coluna_label = ctk.CTkLabel(coluna_frame, text="Coluna:")
coluna_label.pack(side="left", padx=5)

coluna_var = StringVar()
entry_coluna = ctk.CTkEntry(coluna_frame, textvariable=coluna_var, width=60)
entry_coluna.pack(side="left", padx=10)
entry_coluna.insert(0, "A")  # Valor padrão

info_label = ctk.CTkLabel(coluna_frame, text="Ex: A, B, C...", text_color="gray")
info_label.pack(side="left")

# Barra de progresso
progress_frame = ctk.CTkFrame(main_frame)
progress_frame.pack(fill="x", padx=10, pady=(20, 5))

progress_bar = ctk.CTkProgressBar(progress_frame)
progress_bar.pack(fill="x", padx=10, pady=5)
progress_bar.set(0)  # Inicializa com 0%

status_label = ctk.CTkLabel(progress_frame, text="Aguardando...")
status_label.pack(pady=5)

# Botão de iniciar
btn_iniciar = ctk.CTkButton(
    main_frame, 
    text="Iniciar Processamento", 
    command=iniciar_processamento,
    fg_color="#28a745",  # Verde
    hover_color="#218838"
)
btn_iniciar.pack(pady=15)

# Informações de rodapé
footer_label = ctk.CTkLabel(root, text="Desenvolvido pela equipe da GCSUB ", text_color="gray")
footer_label.pack(side="bottom", pady=5)

root.mainloop()

